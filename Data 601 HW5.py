from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
wb1 = load_workbook('week_05_homework_XLSX_openpyxl.xlsx', data_only=True)
worksheet_names = wb1.sheetnames
main_sheet = worksheet_names.index('main')
another_sheet = worksheet_names.index('another')
wb1.active = main_sheet
ws = wb1.active
data = ws.values
cols = next(data)[0:]
data = list(data)
df = pd.DataFrame(data, columns=cols)
df2 = df.rename(columns={df.columns[0] : 'p_id'})
print(df2)
wb1.active = another_sheet
ws1 = wb1.active
data1 = ws1.values
cols1 = next(data1)[0:]
data1 = list(data1)
df1 = pd.DataFrame(data1, columns=cols1)
print(df1)
df3 = df2.merge(df1, how='left', on='p_id')
print(df3)
rows = dataframe_to_rows(df3,index = False, header=True)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         ws.cell(row=r_idx, column=c_idx, value=value)
wb1.save("HW5.xlsx")